#Program: randomemployee.py
#Author: Jake Davis
#Date: 7/30/2018
#Description: Generates an excel file with a random number of employee names and email addresses
#INPUT: excel file listing all employees of a company
#OUTPUT: excel file with a random sample of n rows taken from the all employee file
#USAGE: randomemployee.py <filename> <number of employess to select>


#! python3

#load modules
from openpyxl import *
import random, sys

DEBUG = True

if len(sys.argv) != 3:
    print("Error: please follow usage Rules")
    print("USAGE: randomemployee.py <filename> <number of employess to select>")
    sys.exit()

FILENAME = sys.argv[1]
NUMBER = int(sys.argv[2])

#read in the excel file or throw exception
try:
    wb = load_workbook(FILENAME)
except:
    print("ERROR: No workboook found:" + FILENAME)


#create a new wb
newwb = Workbook()
newsheet = newwb.active
#select random number or rows from INPUT file

sheet = wb.active
rows = sheet.max_row
i=0
alreadyChosen = []

while i < NUMBER:
    randomRow = random.randint(1,rows)
    #if the employee has already been chosen, skip and move to another random selection
    if randomRow in alreadyChosen:
        next
    alreadyChosen.append(randomRow)
    newsheet.cell(i+1,2).value = sheet.cell(randomRow,1).value
    newsheet.cell(i+1,1).value = sheet.cell(randomRow,2).value
    newsheet.cell(i+1,3).value = ''
    newsheet.cell(i+1,4).value = 'Petone'
    i += 1

#write selection to OUTPUT file
newwb.save('newlist.xlsx')
#display error or success messages
